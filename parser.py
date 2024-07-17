# BASE LIBRARY
import csv
import random
import time
import lxml
import os
import json
import re
from dotenv import load_dotenv
# EXEL
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import SheetTitleException
# THREADING
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
import threading
# Multiprocess
# from multiprocessing import Pool
# import multiprocessing
# REQUEST&BS4
import requests
from bs4 import BeautifulSoup
import fake_useragent
# SELENIUM
# from selenium import webdriver
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.chrome.options import Options
# from webdriver_manager.chrome import ChromeDriverManager
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.common.exceptions import StaleElementReferenceException
# from selenium.common.exceptions import TimeoutException
# from selenium.webdriver.common.by import By
# MY IMPORTS
from settings.ParserClass import Parser

URL = 'https://www.belarusinfo.by'   #URL сайта
URL_CAT = 'https://www.belarusinfo.by/ru/company/kompyutery-i-internet'


class Time(Parser):
    def __init__(self):
        self.time_start = time.time()  # Сохранение начального времени

    def end(self):
        time_end = time.time()
        elapsed_time = time_end - self.time_start  # Вычисление затраченного времени
        self.logger(f'|--- Сбор всех данных занял {elapsed_time:.2f} секунд.')  # Логирование времени


parser = Parser()


def parse_categs_on_main(gen_url=URL, parser=parser):
    """
    Парсит категории базовой страницы https://www.belarusinfo.by...
    """
    try:
        soup = BeautifulSoup(parser.fetch_data(gen_url).text, 'lxml')
        urls = soup.find('div', {'id': 'position2'}).find('div', {'class': 'customrubricss'}).find_all('a')
        urls_dict = {}
        # Берём значения из conf.env
        load_dotenv(dotenv_path=parser.file_path_env)
        MAIN_GROUP = os.getenv('MAIN_GROUP')
        parser.file_path_book = os.path.join('data', 'parsed', f'{clear_name_for_book(MAIN_GROUP)}.xlsx')
        for a in urls:
            if a.text == MAIN_GROUP:
                urls_dict[a.text] = gen_url + a.attrs['href'].replace('.html', '')
        # create_new_book(parser.file_path_book)
        return urls_dict
    except Exception as e:
        parser.logger(f'Ошибка работы В функции parse_categs_on_main(gen_url=URL, parser=parser): {e}', saveonly=False, first=False, infunction=True)


def parse_categs(categ_urls_dict, gen_url=URL, parser=parser):
    """
    Парсит подкатегории страницы https://www.belarusinfo.by/ru/, полученные из функции parse_categs_on_main(...)
    """
    try:
        for categ_name, url in categ_urls_dict.items():
            soup = BeautifulSoup(parser.fetch_data(url).text, 'lxml')
            categs_table = soup.find('div', {'id': 'bodycontent'}).find('div', {'class': 'rubricator_container'}).find_all('a')
            urls_dict = {}
            # Берём значения из conf.env
            load_dotenv(dotenv_path=parser.file_path_env)
            ADDITIONAL_GROUPS = os.getenv('ADDITIONAL_GROUPS').split('|')
            for a in categs_table:
                if a.text in ADDITIONAL_GROUPS:
                    urls_dict[a.text] = gen_url + a.attrs['href'].replace('.html', '')
            # # urls = [a['href'] for a in categs_table if a.text == 'Компьютеры - программное обеспечение' or a.text == 'ИТ технологии']
            # Сохраняем
            parser.save_data('Categs' ,'data', urls_dict)
            return urls_dict
    except Exception as e:
        parser.logger(f'Ошибка работы В функции parse_categs(categ_urls_dict, gen_url=URL, parser=parser): {e}', saveonly=False, first=False, infunction=True)


def parse_process(task, lock, url=URL, parser=parser):
    try:
        categ_name, categ_url = task
        # print(f"Name: {categ_name}, category: {categ_url}")

        for page_num in infinite_step_generator(0, 10):
            # Находим объект для проверки на существование компаний на странице пагинатора
            param = f"limitstart={page_num}"
            response_text = parser.fetch_data(url=categ_url, params=param).text
            soup = BeautifulSoup(response_text, 'lxml')
            print(str(categ_url) + '?limitstart=' + str(page_num))

            nothing = soup.find('div', {'id': "zverlistcolumn", "class": "ms_zverlistcolumn_city"}).find('p', {'class': 'nothing'})
            # Условие выхода
            if nothing:
                break

            parse_page(soup, categ_name, page_num, lock)
    except Exception as e:
        parser.logger(f"Ошибка обработки в функции parse_process: {e}")


def clear_name_for_book(name, parser=parser):
    """
    Очищает имя модели для создания книги в Excel.
    Возвращает очищенное имя, которое содержит только английские буквы и цифры и имеет длину не более 30 символов.
    """
    try:
        # Удаление недопустимых символов, оставляя только английские буквы и цифры
        cleaned_name = re.sub(r'[^a-zA-Zа-яА-Я0-9]', ' ', name)
        # Обрезка до 30 символов
        cleaned_name = cleaned_name[:30]
        # Удаление пробелов в начале и в конце
        cleaned_name = cleaned_name.strip()
        return cleaned_name
    except Exception as e:
        parser.logger(f"Ошибка обработки в функции clear_model_name_for_book: {e}")


def infinite_step_generator(start, step):
    """
    Функция для циклического прохода по всем ссылкам (решение обхода пагинации)
    """
    while True:
        yield start
        start += step


def parse_page(soup, categ_name, page_num, lock, parser=parser):
    boxes = soup.find('div', {'id': "zverlistcolumn", "class": "ms_zverlistcolumn_city"}).find_all('div', {'class': 'zvers_c 111'})
    # print(boxes)
    parser.file_path_scv = os.path.join('data', 'parsed')
    if boxes:
        # data = []
        for box in boxes:
            try:
                name = safe_get_text(box, [('div', {'class': 'contentcartochkazver'}), ('div', {'class': 's_c_title'})])
                address = safe_get_text(box, [('div', {'class': 'infozvercartohkalist'}),('div', {'class': 's_c_m adrezz'})])
                phone = ' '.join(safe_get_text(soup, [('div', {'class': 'telephonzz'}), ('span', {'class': 'ch_phone'})]).split())
                site = safe_get_text(box, [('div', {'class': 'saitzzz'})])
                if name and address and phone and site:
                    # print(f"NAME:{name}\nADDRES:{address}\nPHONE:{phone}\nSITE:{site}\n")
                    row = [name, address, phone, site]
                    add_to_scv(parser.file_path_scv, categ_name, row, lock)
                    # data.append(row)

            except Exception as e:
                parser.logger(f'Ошибка работы В функции parse_page(soup, parser=parser) при работе с категорией: "{categ_name}", page_number:{page_num}: {e}',
                              saveonly=False, first=False, infunction=True)


def safe_get_text(parent, selectors, default='-'):
    """
    Безопасное извлечение элементов из страницы и в случае неудачи присвоение им '-'
    """
    try:
        element = parent
        for selector in selectors:
            element = element.find(selector[0], selector[1])
        return element.text.strip()
    except AttributeError:
        return default


def create_new_book(file_path, parser=parser):
    """
    Создает новую книгу и добавляет лист с указанным названием.
    """
    try:
        # Проверка, существует ли уже файл
        if os.path.exists(file_path):
            parser.logger(f"Файл {file_path} уже существует. Перезаписываем файл.")

        # Создание новой книги
        workbook = Workbook()
        sheet = workbook.active

        # Переименование листа
        sheet.title = 'NoneList'

        # Сохранение книги в файл
        workbook.save(file_path)
        parser.logger(f"Книга '{file_path}' с листом 'NoneList' успешно создана.")

    except Exception as e:
        parser.logger(f"Произошла ошибка при создании книги create_new_book: {e}")


def add_to_scv(dir_path, categ_name, row, lock, parser=parser):
    """
    Создаёт или добавляет файлы scv
    """
    columns = ['NAME', 'ADDRESS', 'PHONE', 'SITE']

    # Формируем имя файла и полный путь
    file_name = f'{categ_name.replace(" ", "_")}.csv'
    file_path = os.path.join(dir_path, file_name)

    try:
        with lock:  # Используем блокировку для обеспечения потокобезопасности
            # Открываем файл для добавления данных
            file_exists = os.path.exists(file_path)

            with open(file_path, mode='a', newline='') as file:
                writer = csv.DictWriter(file, fieldnames=columns)

                # Записываем заголовки только если файл был создан только что
                if not file_exists:
                    parser.logger(f"Файл по пути '{file_path}' не найден. Создаем файл.")
                    writer.writeheader()  # Записываем заголовки

                # Записываем данные
                writer.writerow({
                    'NAME': row[0],
                    'ADDRESS': row[1],
                    'PHONE': row[2],
                    'SITE': row[3],
                })
                parser.logger(f"[+] Добавлены данные в файл '{file_path}': {row}")

    except Exception as e:
        parser.logger(f"Произошла ошибка при добавлении данных в файл {file_path}: {e}")


def add_to_sheet(file_path, sheet_name, row, lock, parser=parser):
    try:
        cleared_sheet_name = clear_name_for_book(sheet_name)
        columns = ['NAME', 'ADDRESS', 'PHONE', 'SITE']
        # Проверяем существование файла
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Файл по пути {file_path} не существует.")

        with lock:
            workbook = load_workbook(file_path)

            # Проверяем наличие листа
            if cleared_sheet_name in workbook.sheetnames:
                sheet = workbook[cleared_sheet_name]
            else:
                # Создаём новый лист, если не найден
                sheet = workbook.create_sheet(title=cleared_sheet_name)
                sheet.append(columns)
                parser.logger(f'Лист с названием "{cleared_sheet_name}" был создан в книге по пути "{file_path}"')
            sheet.append(row)
            parser.logger(f"[+] : [ NAME:{row[0]} | ADDRESS:{row[1]} | PHONE:{row[2]} | SITE:{row[3]}] ]")

            # Сохраняем изменения
            workbook.save(file_path)

    except FileNotFoundError as fnf_error:
        parser.logger(f'Ошибка: {fnf_error}')
    except ValueError as val_error:
        parser.logger(f'Ошибка: {val_error}')
    except SheetTitleException as st_error:
        parser.logger(f'Ошибка: Недопустимое имя листа "{cleared_sheet_name}": {st_error}')
    except Exception as e:
        parser.logger(f'Ошибка добавления данных в лист "{cleared_sheet_name}" в функции add_to_sheet: {e}')


def clear_data_parsed(dir_path):
    """
    Очищает все файлы и поддиректории в указанной директории.

    :param dir_path: Путь к директории, которую нужно очистить.
    """
    try:
        # Проверяем, существует ли директория
        if os.path.exists(dir_path):
            # Проходим по всем файлам и поддиректориям в директории
            for item in os.listdir(dir_path):
                item_path = os.path.join(dir_path, item)

                # Если это файл, удаляем его
                if os.path.isfile(item_path):
                    os.remove(item_path)
                    print(f"Удален файл: {item_path}")
        else:
            print(f"Директория по пути '{dir_path}' не существует.")

    except Exception as e:
        print(f"Произошла ошибка при очистке директории '{dir_path}': {e}")

def parse(parser):
    try:
        parser.logger('|---Программа начала свою работу---|', False, True)
        parser.logger('|---Настройка...---|', False, True)
        driver = parser.setup_driver()
        time = Time()
        lock = parser.lock_treads
        parser.file_path_env = os.path.join('settings', 'conf.env')
        clear_data_parsed('data/parsed')

        parse_categs(parse_categs_on_main())
        categs_for_parse = parser.read_data('Categs.json', 'data', 'json')

        # Добавление парсинга раздела в потоковую обработку
        tasks = [(name, categ_url) for name, categ_url in categs_for_parse.items()] #  Создание задач
        THREADS = 12
        lock = Lock()
        with ThreadPoolExecutor(max_workers=THREADS) as executor:
            futures = [executor.submit(parse_process, task, lock) for task in tasks]    # Выполнение в потоке функкции
        for future in futures:
            try:
                future.result() # Возврат результатов
            except Exception as e:
                parser.logger(f"Ошибка потока: {e}")


        time.end()
        driver.quit()
        parser.logger(f'|------------------------------------------------------|', saveonly=False, first=False, infunction=False)
        parser.logger(f'|---Модели были успешно собраны по пути data/BMW/output.xlsx.', saveonly=False, first=False, infunction=False)

    except KeyboardInterrupt:
        parser.logger('\nKeyboardInterrupt')
    except Exception as e:
        parser.logger(f'|---Ошибка в работе программы\n')
    finally:
        driver.quit()
        parser.logger(f'|------------------------------------------------------|', saveonly=False, first=False, infunction=False)
        parser.logger('|---Завершение работы программы...', saveonly=False, first=False, infunction=False)


if __name__ == "__main__":
    parse(parser)